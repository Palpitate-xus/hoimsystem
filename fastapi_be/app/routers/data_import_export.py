import datetime
import io

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import ADMIN_ROLES, User, require_roles
from app.models import Department, Doctor, Patient, Pharmaceutical, User as UserModel
from app.security import hash_password

router = APIRouter()

MAX_IMPORT_SIZE = 10 * 1024 * 1024
ENTITY_HEADERS = {
    "doctors": ["name", "sex", "title", "education", "phone", "department_id", "permission", "username", "password"],
    "patients": ["name", "sex", "identity", "birthday", "phone", "address", "permission", "allergy_history"],
    "pharmaceuticals": ["name", "stock", "price", "expireddate", "supplier", "remark", "antibiotic_level", "status"],
}


def _validate_entity(entity: str) -> str:
    if entity not in ENTITY_HEADERS:
        raise HTTPException(status_code=400, detail="仅支持 doctors、patients、pharmaceuticals")
    return entity


def _as_int(value, field: str, row_number: int, default=None):
    if value in (None, ""):
        return default
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"第{row_number}行{field}必须是整数") from exc


def _as_float(value, field: str, row_number: int, default=None):
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"第{row_number}行{field}必须是数字") from exc


def _as_date(value, field: str, row_number: int):
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    try:
        return datetime.datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"第{row_number}行{field}必须是YYYY-MM-DD") from exc


def _rows_from_workbook(contents: bytes):
    try:
        workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="无法读取Excel文件") from exc
    if not rows:
        raise HTTPException(status_code=400, detail="Excel文件为空")
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    return headers, [dict(zip(headers, row)) for row in rows[1:] if any(value not in (None, "") for value in row)]


def _create_doctor(row, row_number, db: Session):
    if not row.get("name"):
        raise ValueError(f"第{row_number}行医生姓名不能为空")
    department_id = _as_int(row.get("department_id"), "department_id", row_number)
    if department_id and not db.query(Department).filter(Department.department_id == department_id).first():
        raise ValueError(f"第{row_number}行科室不存在")
    username = str(row.get("username") or f"doctor_{row.get('phone') or row_number}").strip()
    if db.query(UserModel).filter(UserModel.username == username).first():
        raise ValueError(f"第{row_number}行账号{username}已存在")
    user = UserModel(username=username, password=hash_password(str(row.get("password") or "123456")), user_role="doctor")
    db.add(user)
    db.flush()
    doctor = Doctor(
        name=str(row["name"]).strip(),
        sex=_as_int(row.get("sex"), "sex", row_number, 1),
        title=row.get("title"),
        education=row.get("education"),
        phone=str(row.get("phone") or ""),
        department_id=department_id,
        permission=row.get("permission") or "doctor",
        user_id=user.user_id,
    )
    db.add(doctor)


def _create_patient(row, row_number, db: Session):
    identity = str(row.get("identity") or "").strip()
    if not row.get("name") or not identity:
        raise ValueError(f"第{row_number}行患者姓名和身份证号不能为空")
    if db.query(Patient).filter(Patient.identity == identity).first():
        raise ValueError(f"第{row_number}行身份证号已存在")
    patient = Patient(
        name=str(row["name"]).strip(),
        sex=_as_int(row.get("sex"), "sex", row_number, 1),
        identity=identity,
        birthday=_as_date(row.get("birthday"), "birthday", row_number),
        phone=str(row.get("phone") or ""),
        address=row.get("address"),
        permission=row.get("permission") or "allow",
        allergy_history=row.get("allergy_history"),
    )
    db.add(patient)
    if not db.query(UserModel).filter(UserModel.username == identity).first():
        db.add(UserModel(username=identity, password=hash_password("123456"), user_role="patient"))


def _create_pharmaceutical(row, row_number, db: Session):
    if not row.get("name"):
        raise ValueError(f"第{row_number}行药品名称不能为空")
    db.add(
        Pharmaceutical(
            name=str(row["name"]).strip(),
            stock=_as_int(row.get("stock"), "stock", row_number, 0),
            price=_as_float(row.get("price"), "price", row_number, 0),
            expireddate=_as_date(row.get("expireddate"), "expireddate", row_number),
            supplier=row.get("supplier"),
            remark=row.get("remark"),
            antibiotic_level=_as_int(row.get("antibiotic_level"), "antibiotic_level", row_number, 0),
            status=_as_int(row.get("status"), "status", row_number, 0),
            purchasing_time=datetime.datetime.now(),
        )
    )


@router.get("/dataImportExport/template/{entity}")
def download_template(entity: str, current_user: User = Depends(require_roles(*ADMIN_ROLES))):
    entity = _validate_entity(entity)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = entity
    sheet.append(ENTITY_HEADERS[entity])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{entity}_template.xlsx"'},
    )


@router.post("/dataImportExport/import/{entity}")
def import_data(entity: str, file: UploadFile = File(...), current_user: User = Depends(require_roles(*ADMIN_ROLES)), db: Session = Depends(get_db)):
    entity = _validate_entity(entity)
    contents = file.file.read()
    if len(contents) > MAX_IMPORT_SIZE:
        raise HTTPException(status_code=400, detail="文件不能超过10MB")
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="仅支持xlsx或xlsm文件")
    headers, rows = _rows_from_workbook(contents)
    missing = [header for header in ENTITY_HEADERS[entity] if header not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"缺少列: {', '.join(missing)}")
    imported = 0
    errors = []
    try:
        for index, row in enumerate(rows, start=2):
            try:
                if entity == "doctors":
                    _create_doctor(row, index, db)
                elif entity == "patients":
                    _create_patient(row, index, db)
                else:
                    _create_pharmaceutical(row, index, db)
                imported += 1
            except ValueError as exc:
                errors.append(str(exc))
        if errors:
            db.rollback()
        else:
            db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail="导入失败，请检查数据后重试") from exc
    return {"code": 200 if not errors else 422, "msg": "导入完成" if not errors else "存在错误，未写入任何数据", "data": {"entity": entity, "total": len(rows), "imported": imported if not errors else 0, "errors": errors}}


def _export_rows(entity: str, db: Session):
    if entity == "doctors":
        return [[item.name, item.sex, item.title, item.education, item.phone, item.department_id, item.permission, item.user.username if item.user else ""] for item in db.query(Doctor).order_by(Doctor.doctor_id).all()]
    if entity == "patients":
        return [[item.name, item.sex, item.identity, item.birthday, item.phone, item.address, item.permission, item.allergy_history] for item in db.query(Patient).order_by(Patient.patient_id).all()]
    return [[item.name, item.stock, item.price, item.expireddate, item.supplier, item.remark, item.antibiotic_level, item.status] for item in db.query(Pharmaceutical).order_by(Pharmaceutical.pharmaceutical_id).all()]


def _sanitize_cell(value):
    """防 Excel 公式注入：以 = + - @ 开头的文本前置单引号（openpyxl 写入 data_type='s'）。"""
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return f"'{value}"
    return value


@router.get("/dataImportExport/export/{entity}")
def export_data(entity: str, current_user: User = Depends(require_roles(*ADMIN_ROLES)), db: Session = Depends(get_db)):
    entity = _validate_entity(entity)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = entity
    sheet.append(ENTITY_HEADERS[entity])
    for row in _export_rows(entity, db):
        sheet.append([_sanitize_cell(v) for v in row])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{entity}.xlsx"'},
    )
