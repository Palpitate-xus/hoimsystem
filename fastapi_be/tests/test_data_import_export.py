import io

import pytest
from openpyxl import Workbook, load_workbook

from app.models import Patient


def xlsx_bytes(headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.mark.asyncio
class TestDataImportExport:
    async def test_admin_imports_patients_and_exports_them(self, async_client, seed_data, auth_headers, db_session):
        headers = ["name", "sex", "identity", "birthday", "phone", "address", "permission", "allergy_history"]
        content = xlsx_bytes(headers, [["导入患者", 1, "110101199901019999", "1999-01-01", "13900009999", "北京", "allow", "青霉素"]])
        response = await async_client.post(
            "/api/dataImportExport/import/patients",
            headers=auth_headers(seed_data["admin_user"].username),
            files={"file": ("patients.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.json()["data"]["imported"] == 1
        assert db_session.query(Patient).filter(Patient.identity == "110101199901019999").count() == 1

        exported = await async_client.get("/api/dataImportExport/export/patients", headers=auth_headers(seed_data["admin_user"].username))
        assert exported.status_code == 200
        workbook = load_workbook(io.BytesIO(exported.content), read_only=True)
        assert workbook.active.cell(1, 1).value == "name"

    async def test_import_rejects_missing_columns_and_patient_forbidden(self, async_client, seed_data, auth_headers):
        content = xlsx_bytes(["name"], [["缺列"]])
        response = await async_client.post(
            "/api/dataImportExport/import/patients",
            headers=auth_headers(seed_data["admin_user"].username),
            files={"file": ("patients.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 400
        forbidden = await async_client.get(
            "/api/dataImportExport/export/patients",
            headers=auth_headers(seed_data["patient_user"].username),
        )
        assert forbidden.status_code == 403
